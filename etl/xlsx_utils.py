"""Utilities for reading xlsx files with openpyxl."""

import os
from datetime import datetime, date
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

DATA_DIR = os.getenv("DATA_DIR", os.path.join(os.path.dirname(__file__), "..", "docs", "event_docs"))


def data_path(filename):
    return os.path.join(DATA_DIR, filename)


def open_workbook(filename):
    """Open a workbook in read-only mode for performance."""
    return load_workbook(data_path(filename), read_only=True, data_only=True)


def sheet_to_dicts(ws, header_row=1):
    """Convert a worksheet to a list of dicts keyed by header names.

    Skips rows that are entirely None.
    """
    rows = ws.iter_rows(min_row=header_row)
    headers = [cell.value for cell in next(rows)]

    # Clean up header names — strip whitespace, replace None with positional name
    clean_headers = []
    for i, h in enumerate(headers):
        if h is None:
            clean_headers.append(f"_col_{i}")
        else:
            clean_headers.append(str(h).strip())
    headers = clean_headers

    results = []
    for row in rows:
        values = [cell.value for cell in row]
        if all(v is None for v in values):
            continue
        results.append(dict(zip(headers, values)))
    return results


def safe_str(val, max_len=None):
    """Convert to string or None, optionally truncating."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if max_len:
        s = s[:max_len]
    return s


def safe_int(val):
    """Convert to int or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return None


def safe_float(val):
    """Convert to float or None."""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).strip().replace(",", ""))
    except (ValueError, TypeError):
        return None


def safe_bool(val):
    """Convert to bool or None. Handles Y/N, Yes/No, True/False, 1/0."""
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ("y", "yes", "true", "1"):
        return True
    if s in ("n", "no", "false", "0"):
        return False
    return None


def safe_date(val):
    """Convert to date or None."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def normalize_gender(val):
    """Normalize gender to single char M/F or None."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s in ("M", "MALE"):
        return "M"
    if s in ("F", "FEMALE"):
        return "F"
    return None


def normalize_food_pref(val):
    """Normalize food preference to code."""
    if val is None:
        return None
    s = str(val).strip().upper()
    mapping = {
        "R": "R", "REGULAR": "R",
        "J": "J", "JAIN": "J",
        "JW": "JW", "JAIN WESTERN": "JW",
        "RW": "RW", "REGULAR WESTERN": "RW",
        "W": "W", "WESTERN": "W",
    }
    return mapping.get(s, s[:10])
